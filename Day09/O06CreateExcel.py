
from openpyxl import  Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "Bangalore"

# cell
ws['B9'] = "Hello World"
ws['D9'] = 1048576
ws['F9'] = datetime.now()
ws['H9'].value = "=column()"

wb.save("FirstExcel.xlsx")
