
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, PieChart3D
from openpyxl.chart.label import DataLabelList

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

data = [
    ('Products', 'Sales'),
    ('Pepsi', 450),
    ('Coke', 380),
    ('Sprite', 300),
    ('Mirinda', 350),
    ('Thumbs up', 490),
    ('Fanta', 275),
    ('Slice', 400)
]

print(f"data :{data}")

for row in data:
    ws.append(row)

dataRef = Reference(ws, min_row= 2, max_row=8, min_col=2)
labelRef = Reference(ws, min_row=2, max_row=8, min_col=1)

chart = BarChart()
chart.add_data(dataRef)
chart.set_categories(labelRef)
chart.title = 'Baverage Sales'
# chart.x_axis = labelRef
# chart.dataLabels = DataLabelList()
# chart.dataLabels.showVal = True

ws.add_chart(chart, "E9")
wb.save('FirstExcel.xlsx')





