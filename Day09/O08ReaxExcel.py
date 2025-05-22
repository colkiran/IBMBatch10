
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from prettytable import PrettyTable

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

print(ws.dimensions)
print('-' * 60)

res = range_boundaries(ws.dimensions)
# min_col, min_row, max_col, max_row
print(res)

print(res[2] - res[0] + 1)

print('-' * 60)
dataRange = ws[ws.dimensions]

row = 1
for c1, c2, c3, c4, c5 in dataRange:
    if row == 1:
        plyrTbl = PrettyTable([c1.value, c2.value, c3.value, c4.value, c5.value])
    else:
        plyrTbl.add_row([c1.value, c2.value, c3.value, c4.value, c5.value])
    row += 1

plyrTbl.align['PlayerName'] = 'l'
plyrTbl.align['Country'] = 'r'


print(plyrTbl)
